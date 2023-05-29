import time
import unittest
import openpyxl # 1. Importante añadir la libreria

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

class Funciones_Excel (): # 2. Se crea la clase Funciones_Excel y se crea la inicialización dentro de esta
    def __init__(self,driver):
        self.driver = driver

    # Función que obtiene la ruta del archivo de Excel y cuenta la cantidad de filas con información
    def getRowCount(self, directorio_ruta, nombre_hoja):
        Workbook = openpyxl.load_workbook(directorio_ruta) # 3. Se especifica la ruta en la que se encuentra el archivo de Excel
        sheet = Workbook[nombre_hoja] # 4. Se indica el nombre de la hoja que contiene los datos
        return (sheet.max_row) # 5. Cuenta el maximo de filas que tiene el documento de Excel

    # Función que obtiene la ruta del archivo de Excel y cuenta la cantidad de columnas con información
    def getColumnCount(self, directorio_ruta, nombre_hoja):
        Workbook = openpyxl.load_workbook(directorio_ruta) # 6. Se especifica la ruta en la que se encuentra el archivo de Excel
        sheet = Workbook[nombre_hoja] # 7. Se indica el nombre de la hoja que contiene los datos
        return (sheet.max_column) # 8. Cuenta el maximo de columnas que tiene el documento de Excel

    # Función que obtiene la ruta del archivo de Excel y lee la información de las filas y las columnas
    def readData(self, directorio_ruta, nombre_hoja, num_filas, num_columnas):
        Workbook = openpyxl.load_workbook(directorio_ruta)  # 9. Se especifica la ruta en la que se encuentra el archivo de Excel
        sheet = Workbook[nombre_hoja]  # 10. Se indica el nombre de la hoja que contiene los datos
        # 11. Obtiene el numero de filas y columnas y lo guarda en las variables row y column
        return sheet.cell(row = num_filas, column = num_columnas).value

    # Función que obtiene la ruta del archivo de Excel y permite escribir el valor de data donde se le especifique
    def writeData(self, directorio_ruta, nombre_hoja, num_filas, num_columnas, data):
        Workbook = openpyxl.load_workbook(directorio_ruta)  # 12. Se especifica la ruta en la que se encuentra el archivo de Excel
        sheet = Workbook[nombre_hoja]  # 13. Se indica el nombre de la hoja que contiene los datos
        sheet.cell(row=num_filas,column=num_columnas).value = data
        Workbook.save(directorio_ruta) # 14. Guarda la información de la variable data en el archivo de Excel
